from flask import Flask, render_template, redirect, url_for, flash, request, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_wtf.file import FileField, FileAllowed, FileSize
from passlib.hash import pbkdf2_sha256
from wtforms import StringField, PasswordField, SelectField, TextAreaField, SubmitField, DateTimeLocalField
from wtforms.validators import DataRequired, Length, Optional, ValidationError
from wtforms_sqlalchemy.fields import QuerySelectField
from cachelib import SimpleCache
from datetime import datetime
import uuid
import re
import openpyxl
import os
from werkzeug.utils import secure_filename
from sqlalchemy.orm import joinedload
from config import Config
from models import db, Admin, Undangan, Tamu, Rekening, Galeri, Cerita, Tema, Category
from dummy import DUMMY_UNDANGAN, DUMMY_TAMU, DUMMY_REKENING_LIST, DUMMY_GALERI_LIST, DUMMY_CERITA_LIST

app = Flask(__name__)
app.config.from_object(Config)
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
db.init_app(app)
csrf = CSRFProtect(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'
cache = SimpleCache()

from datetime import datetime
app.jinja_env.filters['datetimeformat'] = lambda value, format: datetime.now().strftime(format)
app.jinja_env.globals.update(getattr=getattr)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Custom File Size Validator
def file_size(max_size):
    def _file_size(form, field):
        if field.data and hasattr(field.data, 'filename') and field.data.filename:
            field.data.seek(0, os.SEEK_END)
            size = field.data.tell()
            field.data.seek(0)
            if size > max_size:
                raise ValidationError(f'File harus kurang dari {max_size // (1024 * 1024)}MB.')
    return _file_size

# Form Definitions
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

class SuperAdminForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
    password = PasswordField('Password', validators=[Length(min=6, message="Password harus minimal 6 karakter jika diisi.")])
    submit = SubmitField('Simpan')

class NewAdminForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=6)])
    slug = StringField('Slug Undangan', validators=[DataRequired(), Length(min=3, max=50)])
    tema_id = SelectField('Tema', coerce=int, validators=[DataRequired()])
    submit = SubmitField('Simpan')

    def __init__(self, *args, **kwargs):
        super(NewAdminForm, self).__init__(*args, **kwargs)
        self.tema_id.choices = [(tema.id, tema.nama) for tema in Tema.query.all()]

class EditAdminForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
    password = PasswordField('Password', validators=[Length(min=6, message="Password harus minimal 6 karakter jika diisi.")])
    slug = StringField('Slug Undangan', validators=[DataRequired(), Length(min=3, max=50)])
    tema_id = SelectField('Tema', coerce=int, validators=[DataRequired()])
    submit = SubmitField('Simpan')

    def __init__(self, *args, **kwargs):
        super(EditAdminForm, self).__init__(*args, **kwargs)
        self.tema_id.choices = [(tema.id, tema.nama) for tema in Tema.query.all()]

class CategoryForm(FlaskForm):
    name = StringField('Nama Kategori', validators=[DataRequired(), Length(max=50)])
    submit = SubmitField('Simpan')

class TemaForm(FlaskForm):
    nama = StringField('Nama Tema', validators=[DataRequired(), Length(max=50)])
    template_name = StringField('Nama Template', validators=[DataRequired(), Length(max=100)])
    description = TextAreaField('Deskripsi', validators=[Optional()])
    image_url = FileField('Gambar Tema', validators=[
        Optional(),
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    category_id = QuerySelectField('Kategori', query_factory=lambda: Category.query.all(), get_label='name', allow_blank=True, blank_text='Pilih Kategori (Opsional)', validators=[Optional()])
    submit = SubmitField('Simpan Tema')

class UndanganForm(FlaskForm):
    nama_mempelai = StringField('Nama Kedua Mempelai', validators=[DataRequired(), Length(max=200)])
    mempelai_pria = StringField('Mempelai Pria', validators=[DataRequired(), Length(max=100)])
    bio_pria = TextAreaField('Bio Pria', validators=[Optional()])
    ayah_pria = StringField('Ayah Pria', validators=[Optional(), Length(max=100)])
    ibu_pria = StringField('Ibu Pria', validators=[Optional(), Length(max=100)])
    instagram_pria = StringField('Instagram Pria', validators=[Optional(), Length(max=100)])
    mempelai_wanita = StringField('Mempelai Wanita', validators=[DataRequired(), Length(max=100)])
    bio_wanita = TextAreaField('Bio Wanita', validators=[Optional()])
    ayah_wanita = StringField('Ayah Wanita', validators=[Optional(), Length(max=100)])
    ibu_wanita = StringField('Ibu Wanita', validators=[Optional(), Length(max=100)])
    instagram_wanita = StringField('Instagram Wanita', validators=[Optional(), Length(max=100)])
    tanggal_akad = DateTimeLocalField('Tanggal Akad', format='%Y-%m-%dT%H:%M', validators=[Optional()])
    tempat_akad = StringField('Tempat Akad', validators=[Optional(), Length(max=100)])
    lokasi_akad = TextAreaField('Lokasi Akad', validators=[Optional()])
    maps_akad = TextAreaField('Link Google Maps Akad', validators=[Optional()])
    tanggal_resepsi = DateTimeLocalField('Tanggal Resepsi', format='%Y-%m-%dT%H:%M', validators=[Optional()])
    tempat_resepsi = StringField('Tempat Resepsi', validators=[Optional(), Length(max=100)])
    lokasi_resepsi = TextAreaField('Lokasi Resepsi', validators=[Optional()])
    maps_resepsi = TextAreaField('Link Google Maps Resepsi', validators=[Optional()])
    penerima_kado = StringField('Penerima Kado', validators=[Optional(), Length(max=100)])
    alamat_kado = TextAreaField('Alamat Kado', validators=[Optional()])
    wa = StringField('Nomor WhatsApp', validators=[Optional(), Length(max=20)])
    foto_pria = FileField('Foto Pria', validators=[
        Optional(),
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    foto_wanita = FileField('Foto Wanita', validators=[
        Optional(),
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    audio = FileField('Audio Latar', validators=[
        Optional(),
        FileAllowed(['mp3'], 'Hanya file MP3 yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    bg_sampul = FileField('Background Sampul', validators=[
        Optional(),
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    bg_undangan = FileField('Background Undangan', validators=[
        Optional(),
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    submit = SubmitField('Simpan')

class RekeningForm(FlaskForm):
    nama_bank = StringField('Nama Bank', validators=[DataRequired(), Length(max=50)])
    nomer_rekening = StringField('Nomor Rekening', validators=[DataRequired(), Length(max=50)])
    atas_nama = StringField('Atas Nama', validators=[DataRequired(), Length(max=100)])
    submit = SubmitField('Simpan')

class GaleriForm(FlaskForm):
    url = FileField('Foto', validators=[
        FileAllowed(['jpg', 'jpeg', 'png'], 'Hanya file JPG/PNG yang diizinkan!'),
        FileSize(max_size=16 * 1024 * 1024, message='File harus kurang dari 16MB.')
    ])
    alt = StringField('Deskripsi Foto', validators=[Optional(), Length(max=200)])
    submit = SubmitField('Simpan')

class CeritaForm(FlaskForm):
    judul = StringField('Judul Cerita', validators=[DataRequired(), Length(max=100)])
    tanggal = DateTimeLocalField('Tanggal Cerita', format='%Y-%m-%dT%H:%M', validators=[Optional()])
    isi = TextAreaField('Isi Cerita', validators=[DataRequired()])
    submit = SubmitField('Simpan')

class TamuForm(FlaskForm):
    nama = StringField('Nama Tamu', validators=[DataRequired()])
    submit = SubmitField('Simpan')

class RSVPForm(FlaskForm):
    rsvp_status = SelectField('Status RSVP', choices=[
        ('Hadir', 'Hadir'),
        ('Tidak Hadir', 'Tidak Hadir'),
        ('Masih Ragu', 'Masih Ragu')
    ], validators=[DataRequired()])
    ucapan = TextAreaField('Ucapan', validators=[Optional()])
    submit = SubmitField('Kirim RSVP')

# User Loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Admin, int(user_id))

# Anti-Brute-Force Login
def check_login_attempts(ip):
    attempts = cache.get(f"login_attempts_{ip}") or 0
    if attempts >= 3:
        return False
    cache.set(f"login_attempts_{ip}", attempts + 1, timeout=3600)
    return True

# File Upload Helper
def save_file(file, identifier, field_name):
    if not file or not hasattr(file, 'filename') or not file.filename:
        return None
    try:
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        if file_size > 16 * 1024 * 1024:
            return None
        file.seek(0)
        ext = os.path.splitext(file.filename)[1].lower()
        allowed_extensions = {'.jpg', '.jpeg', '.png'} if field_name in ['foto_pria', 'foto_wanita', 'bg_sampul', 'bg_undangan', 'image_url', 'url'] else {'.mp3'}
        if ext not in allowed_extensions:
            return None
        if field_name == 'image_url':
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'tema')
            filename = secure_filename(f"tema_{identifier}{ext}")
        else:
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], identifier)
            filename = secure_filename(f"{identifier}_{field_name}{ext}")
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        return os.path.join('uploads', 'tema' if field_name == 'image_url' else identifier, filename)
    except Exception as e:
        print(f"Error saving file {field_name}: {str(e)}")
        return None

# Routes
@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/admin')
def admin():
    if not current_user.is_authenticated:
        return redirect(url_for('admin_login'))
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    return redirect(url_for('dashboard'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('superadmin' if current_user.is_superadmin else 'dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        if not check_login_attempts(request.remote_addr):
            flash('Terlalu banyak percobaan login. Coba lagi nanti.', 'error')
            return render_template('admin/login.html', form=form)
        
        admin = Admin.query.filter_by(username=form.username.data).first()
        if admin and pbkdf2_sha256.verify(form.password.data, admin.password_hash):
            login_user(admin)
            cache.delete(f"login_attempts_{request.remote_addr}")
            flash('Login berhasil!', 'success')
            return redirect(url_for('superadmin' if admin.is_superadmin else 'dashboard'))
        flash('Username atau password salah.', 'error')
    return render_template('admin/login.html', form=form)

@app.route('/admin/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin/superadmin')
@login_required
def superadmin():
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    admins = Admin.query.filter(Admin.is_superadmin == 0).options(joinedload(Admin.undangan)).all()
    temas = Tema.query.options(joinedload(Tema.category)).all()
    categories = Category.query.all()
    return render_template('admin/superadmin.html', admins=admins, temas=temas, categories=categories,
                          form=EditAdminForm(), tema_form=TemaForm(), superadmin_form=SuperAdminForm(),
                          category_form=CategoryForm())

@app.route('/admin/superadmin/edit-superadmin', methods=['POST'])
@login_required
def edit_superadmin():
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    form = SuperAdminForm()
    if form.validate_on_submit():
        existing_admin = Admin.query.filter(Admin.username == form.username.data, Admin.id != current_user.id).first()
        if existing_admin:
            flash('Username sudah digunakan.', 'error')
            return redirect(url_for('superadmin'))
        
        current_user.username = form.username.data
        if form.password.data:
            current_user.password_hash = pbkdf2_sha256.hash(form.password.data)
        db.session.commit()
        flash('Akun superadmin berhasil diperbarui!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('superadmin'))

@app.route('/admin/superadmin/new', methods=['GET', 'POST'])
@login_required
def create_admin():
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    form = NewAdminForm()
    if form.validate_on_submit():
        if not re.match(r'^[a-zA-Z0-9-]+$', form.slug.data):
            flash('Slug hanya boleh berisi huruf, angka, dan tanda hubung.', 'error')
            return render_template('admin/form_admin_baru.html', form=form)
        
        if Undangan.query.filter_by(slug=form.slug.data).first():
            flash('Slug sudah digunakan.', 'error')
            return render_template('admin/form_admin_baru.html', form=form)
        
        if Admin.query.filter_by(username=form.username.data).first():
            flash('Username sudah digunakan.', 'error')
            return render_template('admin/form_admin_baru.html', form=form)
        
        undangan = Undangan(
            slug=form.slug.data,
            tema_id=form.tema_id.data,
            nama_mempelai='Admin & Mimin',
            kode=str(uuid.uuid4())[:8]  # Generate unique kode
        )
        db.session.add(undangan)
        db.session.commit()
        
        admin = Admin(
            username=form.username.data,
            password_hash=pbkdf2_sha256.hash(form.password.data),
            is_superadmin=0,
            undangan_id=undangan.id
        )
        db.session.add(admin)
        db.session.commit()
        
        flash('Admin dan undangan berhasil dibuat!', 'success')
        return redirect(url_for('superadmin'))
    
    return render_template('admin/form_admin_baru.html', form=form)

@app.route('/admin/superadmin/edit/<int:admin_id>', methods=['POST'])
@login_required
def edit_admin(admin_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    admin = db.session.get(Admin, admin_id)
    if not admin:
        flash('Admin tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    if admin.is_superadmin:
        flash('Tidak dapat mengedit akun superadmin.', 'error')
        return redirect(url_for('superadmin'))
    
    form = EditAdminForm()
    if form.validate_on_submit():
        if not re.match(r'^[a-zA-Z0-9-]+$', form.slug.data):
            flash('Slug hanya boleh berisi huruf, angka, dan tanda hubung.', 'error')
            return redirect(url_for('superadmin'))
        
        existing_admin = Admin.query.filter(Admin.username == form.username.data, Admin.id != admin_id).first()
        if existing_admin:
            flash('Username sudah digunakan.', 'error')
            return redirect(url_for('superadmin'))
        
        existing_undangan = Undangan.query.filter(Undangan.slug == form.slug.data, Undangan.id != admin.undangan_id).first()
        if existing_undangan:
            flash('Slug sudah digunakan.', 'error')
            return redirect(url_for('superadmin'))
        
        admin.username = form.username.data
        if form.password.data:
            admin.password_hash = pbkdf2_sha256.hash(form.password.data)
        
        undangan = db.session.get(Undangan, admin.undangan_id)
        if undangan:
            undangan.slug = form.slug.data
            undangan.tema_id = form.tema_id.data
        else:
            undangan = Undangan(
                slug=form.slug.data,
                tema_id=form.tema_id.data,
                nama_mempelai='Admin & Mimin',
                kode=str(uuid.uuid4())[:8]
            )
            db.session.add(undangan)
            db.session.commit()
            admin.undangan_id = undangan.id
        
        db.session.commit()
        flash('Admin berhasil diperbarui!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('superadmin'))

@app.route('/admin/superadmin/delete/<int:admin_id>')
@login_required
def delete_admin(admin_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    admin = db.session.get(Admin, admin_id)
    if not admin:
        flash('Admin tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    if admin.is_superadmin:
        flash('Tidak dapat menghapus akun superadmin.', 'error')
        return redirect(url_for('superadmin'))
    
    if admin.id == current_user.id:
        flash('Anda tidak dapat menghapus akun Anda sendiri.', 'error')
        return redirect(url_for('superadmin'))
    
    undangan = db.session.get(Undangan, admin.undangan_id)
    Tamu.query.filter_by(undangan_id=admin.undangan_id).delete()
    db.session.delete(admin)
    if undangan:
        db.session.delete(undangan)
    db.session.commit()
    flash('Admin dan undangan terkait berhasil dihapus.', 'success')
    return redirect(url_for('superadmin'))

@app.route('/admin/category/new', methods=['GET', 'POST'])
@login_required
def create_category():
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    form = CategoryForm()
    if form.validate_on_submit():
        if Category.query.filter_by(name=form.name.data).first():
            flash('Nama kategori sudah digunakan.', 'error')
            return render_template('admin/form_category_baru.html', form=form)
        
        category = Category(name=form.name.data)
        db.session.add(category)
        db.session.commit()
        flash(f'Kategori {category.name} berhasil dibuat!', 'success')
        return redirect(url_for('superadmin'))
    
    return render_template('admin/form_category_baru.html', form=form)

@app.route('/admin/category/edit/<int:category_id>', methods=['POST'])
@login_required
def edit_category(category_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    category = db.session.get(Category, category_id)
    if not category:
        flash('Kategori tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    
    form = CategoryForm()
    if form.validate_on_submit():
        existing_category = Category.query.filter(Category.name == form.name.data, Category.id != category_id).first()
        if existing_category:
            flash('Nama kategori sudah digunakan.', 'error')
            return redirect(url_for('superadmin'))
        
        category.name = form.name.data
        db.session.commit()
        flash(f'Kategori {category.name} berhasil diperbarui!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('superadmin'))

@app.route('/admin/category/delete/<int:category_id>')
@login_required
def delete_category(category_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    category = db.session.get(Category, category_id)
    if not category:
        flash('Kategori tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    
    if Tema.query.filter_by(category_id=category.id).first():
        flash('Kategori tidak dapat dihapus karena masih digunakan oleh tema.', 'error')
        return redirect(url_for('superadmin'))
    
    db.session.delete(category)
    db.session.commit()
    flash(f'Kategori {category.name} berhasil dihapus.', 'success')
    return redirect(url_for('superadmin'))

@app.route('/admin/tema/baru', methods=['GET', 'POST'])
@login_required
def create_tema():
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    form = TemaForm()
    if form.validate_on_submit():
        try:
            tema = Tema(
                nama=form.nama.data,
                template_name=form.template_name.data,
                description=form.description.data,
                category_id=form.category_id.data.id if form.category_id.data else None
            )
            db.session.add(tema)
            db.session.flush()
            if form.image_url.data and hasattr(form.image_url.data, 'filename') and form.image_url.data.filename:
                new_file = save_file(form.image_url.data, tema.id, 'image_url')
                if new_file:
                    tema.image_url = new_file
                else:
                    flash('Gagal mengunggah gambar tema. Pastikan file valid dan kurang dari 16MB.', 'error')
            db.session.commit()
            flash('Tema berhasil ditambahkan!', 'success')
            return redirect(url_for('superadmin'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan tema: {str(e)}', 'error')
            print(f"Error creating tema: {str(e)}")
    
    return render_template('admin/form_tema_baru.html', form=form)

@app.route('/admin/tema/edit/<int:tema_id>', methods=['GET', 'POST'])
@login_required
def edit_tema(tema_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    tema = db.session.get(Tema, tema_id)
    if not tema:
        flash('Tema tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    
    form = TemaForm(obj=tema)
    if form.validate_on_submit():
        try:
            tema.nama = form.nama.data
            tema.template_name = form.template_name.data
            tema.description = form.description.data
            tema.category_id = form.category_id.data.id if form.category_id.data else None
            if form.image_url.data and hasattr(form.image_url.data, 'filename') and form.image_url.data.filename:
                new_file = save_file(form.image_url.data, tema.id, 'image_url')
                if new_file:
                    tema.image_url = new_file
                else:
                    flash('Gagal mengunggah gambar tema. Pastikan file valid dan kurang dari 16MB.', 'error')
            db.session.commit()
            flash('Tema berhasil diperbarui!', 'success')
            return redirect(url_for('superadmin'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui tema: {str(e)}', 'error')
            print(f"Error updating tema: {str(e)}")
    
    return render_template('admin/form_tema_baru.html', form=form, tema=tema)

@app.route('/admin/tema/delete/<int:tema_id>')
@login_required
def delete_tema(tema_id):
    if not current_user.is_superadmin:
        flash('Akses ditolak. Hanya superadmin yang dapat mengakses halaman ini.', 'error')
        return redirect(url_for('dashboard'))
    
    tema = db.session.get(Tema, tema_id)
    if not tema:
        flash('Tema tidak ditemukan.', 'error')
        return redirect(url_for('superadmin'))
    
    if Undangan.query.filter_by(tema_id=tema.id).first():
        flash('Tema tidak dapat dihapus karena masih digunakan oleh undangan.', 'error')
        return redirect(url_for('superadmin'))
    
    db.session.delete(tema)
    db.session.commit()
    flash(f'Tema {tema.nama} berhasil dihapus.', 'success')
    return redirect(url_for('superadmin'))

@app.route('/admin/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    undangan = db.session.get(Undangan, current_user.undangan_id)
    if not undangan:
        flash('Undangan tidak ditemukan.', 'error')
        return redirect(url_for('logout'))
    
    form = UndanganForm(obj=undangan)
    tamu_list = Tamu.query.filter_by(undangan_id=current_user.undangan_id).all()
    rekening_list = Rekening.query.filter_by(undangan_id=current_user.undangan_id).all()
    galeri_list = Galeri.query.filter_by(undangan_id=current_user.undangan_id).all()
    cerita_list = Cerita.query.filter_by(undangan_id=current_user.undangan_id).order_by(Cerita.tanggal).all()
    
    total_tamu = len(tamu_list)
    hadir = sum(1 for tamu in tamu_list if tamu.rsvp_status == 'Hadir')
    tidak_hadir = sum(1 for tamu in tamu_list if tamu.rsvp_status == 'Tidak Hadir')
    masih_ragu = sum(1 for tamu in tamu_list if tamu.rsvp_status == 'Masih Ragu')
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            undangan.nama_mempelai = form.nama_mempelai.data
            undangan.mempelai_pria = form.mempelai_pria.data
            undangan.bio_pria = form.bio_pria.data
            undangan.ayah_pria = form.ayah_pria.data
            undangan.ibu_pria = form.ibu_pria.data
            undangan.instagram_pria = form.instagram_pria.data
            undangan.mempelai_wanita = form.mempelai_wanita.data
            undangan.bio_wanita = form.bio_wanita.data
            undangan.ayah_wanita = form.ayah_wanita.data
            undangan.ibu_wanita = form.ibu_wanita.data
            undangan.instagram_wanita = form.instagram_wanita.data
            undangan.tanggal_akad = form.tanggal_akad.data
            undangan.tempat_akad = form.tempat_akad.data
            undangan.lokasi_akad = form.lokasi_akad.data
            undangan.maps_akad = form.maps_akad.data
            undangan.tanggal_resepsi = form.tanggal_resepsi.data
            undangan.tempat_resepsi = form.tempat_resepsi.data
            undangan.lokasi_resepsi = form.lokasi_resepsi.data
            undangan.maps_resepsi = form.maps_resepsi.data
            undangan.penerima_kado = form.penerima_kado.data
            undangan.alamat_kado = form.alamat_kado.data
            undangan.wa = form.wa.data
            
            for field, field_name in [
                (form.foto_pria, 'foto_pria'),
                (form.foto_wanita, 'foto_wanita'),
                (form.audio, 'audio'),
                (form.bg_sampul, 'bg_sampul'),
                (form.bg_undangan, 'bg_undangan')
            ]:
                if field.data and hasattr(field.data, 'filename') and field.data.filename:
                    new_file = save_file(field.data, undangan.slug, field_name)
                    if new_file:
                        setattr(undangan, field_name, new_file)
                    else:
                        flash(f'Gagal mengunggah {field_name}. Pastikan file valid dan kurang dari 16MB.', 'error')
            
            db.session.commit()
            flash('Data undangan berhasil diperbarui!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan data undangan: {str(e)}', 'error')
            print(f"Error saving undangan: {str(e)}")
    elif request.method == 'POST':
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    
    return render_template('admin/dashboard.html', undangan=undangan, tamu_list=tamu_list, rekening_list=rekening_list,
                          galeri_list=galeri_list, cerita_list=cerita_list, form=form, form_tamu=TamuForm(),
                          form_rekening=RekeningForm(), form_galeri=GaleriForm(), form_cerita=CeritaForm(),
                          total_tamu=total_tamu, hadir=hadir, tidak_hadir=tidak_hadir, masih_ragu=masih_ragu,
                          default_tab='tamu')

@app.route('/admin/tamu/new', methods=['POST'])
@login_required
def tambah_tamu():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    form = TamuForm()
    if form.validate_on_submit():
        tamu = Tamu(
            nama=form.nama.data,
            kode=str(uuid.uuid4())[:8],
            undangan_id=current_user.undangan_id
        )
        db.session.add(tamu)
        db.session.commit()
        flash('Tamu berhasil ditambahkan!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/tamu/edit/<int:tamu_id>', methods=['POST'])
@login_required
def edit_tamu(tamu_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    tamu = db.session.get(Tamu, tamu_id)
    if not tamu or tamu.undangan_id != current_user.undangan_id:
        flash('Tamu tidak ditemukan.', 'error')
        return redirect(url_for('dashboard'))
    
    form = TamuForm()
    if form.validate_on_submit():
        tamu.nama = form.nama.data
        db.session.commit()
        flash(f'Tamu {tamu.nama} berhasil diperbarui.', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/admin/tamu/delete/<int:tamu_id>')
@login_required
def delete_tamu(tamu_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    tamu = db.session.get(Tamu, tamu_id)
    if tamu and tamu.undangan_id == current_user.undangan_id:
        db.session.delete(tamu)
        db.session.commit()
        flash('Tamu berhasil dihapus.', 'success')
    else:
        flash('Tamu tidak ditemukan.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/tamu/import', methods=['POST'])
@login_required
def import_tamu():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    if 'file' not in request.files:
        flash('Tidak ada file yang diunggah.', 'error')
        return redirect(url_for('dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Tidak ada file yang dipilih.', 'error')
        return redirect(url_for('dashboard'))
    
    if not file.filename.endswith('.xlsx'):
        flash('File harus berformat .xlsx.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nama = str(row[0]).strip()
            if nama:
                tamu = Tamu(
                    nama=nama,
                    kode=str(uuid.uuid4())[:8],
                    undangan_id=current_user.undangan_id
                )
                db.session.add(tamu)
        db.session.commit()
        flash('Tamu berhasil diimpor!', 'success')
    except Exception as e:
        flash(f'Gagal mengimpor tamu: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/admin/tamu/export')
@login_required
def export_tamu():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    tamu_list = Tamu.query.filter_by(undangan_id=current_user.undangan_id).all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Daftar Tamu'
    sheet.append(['Nama', 'Kode', 'RSVP', 'Ucapan', 'Waktu RSVP'])
    
    for tamu in tamu_list:
        sheet.append([
            tamu.nama,
            tamu.kode,
            tamu.rsvp_status or '',
            tamu.ucapan or '',
            tamu.waktu_rsvp or ''
        ])
    
    filename = f"tamu_undangan_{current_user.undangan_id}.xlsx"
    workbook.save(filename)
    return send_file(filename, as_attachment=True)

@app.route('/admin/rekening/new', methods=['POST'])
@login_required
def tambah_rekening():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    form = RekeningForm()
    if form.validate_on_submit():
        rekening = Rekening(
            nama_bank=form.nama_bank.data,
            nomer_rekening=form.nomer_rekening.data,
            atas_nama=form.atas_nama.data,
            undangan_id=current_user.undangan_id
        )
        db.session.add(rekening)
        db.session.commit()
        flash('Rekening berhasil ditambahkan!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/rekening/delete/<int:rekening_id>')
@login_required
def delete_rekening(rekening_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    rekening = db.session.get(Rekening, rekening_id)
    if rekening and rekening.undangan_id == current_user.undangan_id:
        db.session.delete(rekening)
        db.session.commit()
        flash('Rekening berhasil dihapus.', 'success')
    else:
        flash('Rekening tidak ditemukan.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/galeri/new', methods=['POST'])
@login_required
def tambah_galeri():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    form = GaleriForm()
    if form.validate_on_submit():
        url = save_file(form.url.data, current_user.undangan.slug, 'url')
        if url:
            galeri = Galeri(
                url=url,
                alt=form.alt.data,
                undangan_id=current_user.undangan_id
            )
            db.session.add(galeri)
            db.session.commit()
            flash('Foto berhasil ditambahkan!', 'success')
        else:
            flash('Gagal mengunggah foto. Pastikan file valid dan kurang dari 16MB.', 'error')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/galeri/delete/<int:galeri_id>')
@login_required
def delete_galeri(galeri_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    galeri = db.session.get(Galeri, galeri_id)
    if galeri and galeri.undangan_id == current_user.undangan_id:
        db.session.delete(galeri)
        db.session.commit()
        flash('Foto berhasil dihapus.', 'success')
    else:
        flash('Foto tidak ditemukan.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/cerita/new', methods=['POST'])
@login_required
def tambah_cerita():
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    form = CeritaForm()
    if form.validate_on_submit():
        cerita = Cerita(
            judul=form.judul.data,
            tanggal=form.tanggal.data,
            isi=form.isi.data,
            undangan_id=current_user.undangan_id
        )
        db.session.add(cerita)
        db.session.commit()
        flash('Cerita berhasil ditambahkan!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error di {field}: {error}', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/cerita/delete/<int:cerita_id>')
@login_required
def delete_cerita(cerita_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    cerita = db.session.get(Cerita, cerita_id)
    if cerita and cerita.undangan_id == current_user.undangan_id:
        db.session.delete(cerita)
        db.session.commit()
        flash('Cerita berhasil dihapus.', 'success')
    else:
        flash('Cerita tidak ditemukan.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/admin/ucapan/delete/<int:tamu_id>')
@login_required
def delete_ucapan(tamu_id):
    if current_user.is_superadmin:
        return redirect(url_for('superadmin'))
    
    tamu = db.session.get(Tamu, tamu_id)
    if tamu and tamu.undangan_id == current_user.undangan_id:
        tamu.rsvp_status = None
        tamu.ucapan = None
        tamu.waktu_rsvp = None
        db.session.commit()
        flash('Ucapan berhasil dihapus.', 'success')
    else:
        flash('Ucapan tidak ditemukan.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/katalog')
def katalog():
    category_id = request.args.get('category', type=int)
    if category_id:
        temas = Tema.query.filter_by(category_id=category_id).options(joinedload(Tema.category)).all()
    else:
        temas = Tema.query.options(joinedload(Tema.category)).all()
    categories = Category.query.all()
    return render_template('katalog.html', temas=temas, categories=categories)

@app.route('/katalog/<template_name>')
def preview_tema(template_name):
    tema = Tema.query.filter_by(template_name=template_name).first()
    if not tema:
        flash('Tema tidak ditemukan.', 'error')
        return redirect(url_for('katalog'))
    
    return render_template(f'undangan/{template_name}', 
                         undangan=DUMMY_UNDANGAN,
                         tamu=DUMMY_TAMU,
                         form=RSVPForm(),
                         rekening_list=DUMMY_REKENING_LIST,
                         galeri_list=DUMMY_GALERI_LIST,
                         cerita_list=DUMMY_CERITA_LIST)

@app.route('/<slug>/<kode>', methods=['GET', 'POST'])
def undangan(slug, kode):
    undangan = Undangan.query.filter_by(slug=slug).options(joinedload(Undangan.tema)).first()
    if not undangan:
        flash('Undangan tidak ditemukan.', 'error')
        return redirect(url_for('katalog'))
    
    tamu = Tamu.query.filter_by(kode=kode, undangan_id=undangan.id).first()
    if not tamu:
        flash('Kode tamu tidak valid.', 'error')
        return redirect(url_for('katalog'))
    
    rekening_list = Rekening.query.filter_by(undangan_id=undangan.id).all()
    galeri_list = Galeri.query.filter_by(undangan_id=undangan.id).all()
    cerita_list = Cerita.query.filter_by(undangan_id=undangan.id).order_by(Cerita.tanggal).all()
    form = RSVPForm()
    
    if form.validate_on_submit():
        if tamu.rsvp_status:
            flash('Anda sudah mengisi RSVP.', 'info')
        else:
            tamu.rsvp_status = form.rsvp_status.data
            tamu.ucapan = form.ucapan.data
            tamu.waktu_rsvp = datetime.now()
            db.session.commit()
            flash('RSVP berhasil dikirim!', 'success')
    
    return render_template(f'undangan/{undangan.tema.template_name}', undangan=undangan, tamu=tamu, form=form,
                          rekening_list=rekening_list, galeri_list=galeri_list, cerita_list=cerita_list)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not Admin.query.filter_by(is_superadmin=1).first():
            superadmin = Admin(
                username='superadmin',
                password_hash=pbkdf2_sha256.hash('superadmin123'),
                is_superadmin=1
            )
            db.session.add(superadmin)
            db.session.commit()
    app.run(debug=True)